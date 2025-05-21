from peft import PeftModel
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig, pipeline
import torch
import os
from glob import glob
from transformers import pipeline
from openpyxl import Workbook, load_workbook
import torch
import re
print(torch.__version__)
print(torch.cuda.is_available())  # Should be True
print("CUDA version (build):", torch.version.cuda)
print(torch.cuda.get_device_name(0))  # Should show your GPU

def sanitize(text):
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

def clean_disassembly(text):
    # Remove extra disassembler decoration
    lines = text.splitlines()
    cleaned = []
    for line in lines:
        if re.match(r'^\s*[;\\/|]', line): continue  # skip decorations
        if 'XREF' in line or 'section size' in line: continue
        cleaned.append(line.strip())
    return '\n'.join(cleaned)


# === BitsAndBytes config ===
bnb = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_use_double_quant=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_compute_dtype=torch.float16
)

# === Load base model
base_model = AutoModelForCausalLM.from_pretrained(
    r"D:\LLAMA7code",
    device_map="auto",
    quantization_config=bnb,
    trust_remote_code=True
)

# === Load LoRA adapter
model = PeftModel.from_pretrained(base_model, "lora-checkpoint-ll14")  # or whichever checkpoint dir you saved

# === Load tokenizer
tokenizer = AutoTokenizer.from_pretrained(r"D:\LLAMA7code", use_fast=False)

# Patch pad token if needed
if tokenizer.pad_token is None:
    tokenizer.pad_token = tokenizer.eos_token

# === Use pipeline
pipe = pipeline("text-generation", model=model, tokenizer=tokenizer)

input_dir = "./inputf3"
excel_file = "analysis_resultsFTCLlama.xlsx"


# === Excel setup ===
if os.path.exists(excel_file):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        print(f"[!] Corrupted Excel file. Re-creating it. Reason: {e}")
        os.remove(excel_file)
        wb = Workbook()
        ws = wb.active
        ws.append(["Filename", "Assembly Code (truncated)", "Analysis Result"])
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Filename", "Assembly Code (truncated)", "Analysis Result"])


# === Process all .txt files ===
for filepath in sorted(glob(os.path.join(input_dir, "*.txt"))):
    filename = os.path.basename(filepath)
    
    with open(filepath, "r", encoding="utf-8") as f:
        asm_code = f.read().strip()
        #asm_code = clean_disassembly(asm_code)

    if not asm_code:
        print(f"[!] Skipping empty file: {filename}")
        continue

    prompt = (
    "Analyze this assembly code and tell me if it is malicious or benign. Explain your reasoning.\n\n"
    "```\n"
    f"### Disassembled Code:\n{asm_code.strip()}\n\n"
    "```"
    )

    print(f"[DEBUG] Prompt length (tokens): {len(pipe.tokenizer(prompt)['input_ids'])}")


    print(f"[>] Running analysis on {filename}...")

    output = pipe(prompt, max_new_tokens=512, min_new_tokens=100, do_sample=True, temperature=0.3)
    result_text = output[0]["generated_text"]
    if "### Label:" in result_text:
        result_text = result_text.split("### Label:")[0].strip() + "\n### Label: " + result_text.split("### Label:")[1].strip().split('\n')[0]

    print(result_text)
    ws.append([
    filename,
    sanitize(asm_code[:300] + "..." if len(asm_code) > 300 else asm_code),
    sanitize(result_text.strip())
    ])
    wb.save(excel_file)
    print(f"[✓] Saved result for {filename}")

print(f"\n✅ All results written to: {excel_file}")
