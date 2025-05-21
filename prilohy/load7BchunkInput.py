from peft import PeftModel
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig, pipeline
import torch
import os
from glob import glob
from transformers import pipeline
from openpyxl import Workbook, load_workbook
import re

print(torch.__version__)
print(torch.cuda.is_available())  # Should be True
print("CUDA version (build):", torch.version.cuda)
print(torch.cuda.get_device_name(0))  # Should show your GPU

# === Helpers ===
def sanitize(text):
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

def chunk_code(text, chunk_size=20):
    lines = text.strip().splitlines()
    for i in range(0, len(lines), chunk_size):
        yield "\n".join(lines[i:i + chunk_size])

def extract_first_label(generated_text):
    match = re.search(r"(?i)label\s*:\s*(malicious|benign)", generated_text)
    return match.group(1).lower() if match else "unknown"

# === Config ===
CHUNK_SIZE = 20
MALICIOUS_THRESHOLD = 0.5
input_dir = "./inputf3"
excel_file = "analysis_resultsnewnew.xlsx"

# === BitsAndBytes config ===
bnb = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_use_double_quant=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_compute_dtype=torch.float16
)

# === Load model and tokenizer ===
base_model = AutoModelForCausalLM.from_pretrained(
    r"D:\MistraGood\Mistra",
    device_map="auto",
    quantization_config=bnb,
    trust_remote_code=True
)

model = PeftModel.from_pretrained(base_model, "lora-checkpoint-MistralN3epoch")
tokenizer = AutoTokenizer.from_pretrained(r"D:\MistraGood\Mistra", use_fast=False)
tokenizer.pad_token = tokenizer.pad_token or tokenizer.eos_token
tokenizer.padding_side = "left"
pipe = pipeline("text-generation", model=model, tokenizer=tokenizer)

# === Excel setup ===
if os.path.exists(excel_file):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        print(f"[!] Corrupted Excel file. Re-creating it. Reason: {e}")
        os.remove(excel_file)
        wb = Workbook()
        ws = wb.active
        ws.append(["Filename", "Assembly Code (truncated)", "Analysis Result", "Model Summary"])
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Filename", "Assembly Code (truncated)", "Analysis Result", "Model Summary"])

# === Process all files ===
for filepath in sorted(glob(os.path.join(input_dir, "*.txt"))):
    filename = os.path.basename(filepath)
    with open(filepath, "r", encoding="utf-8") as f:
        asm_code = f.read().strip()

    if not asm_code:
        print(f"[!] Skipping empty file: {filename}")
        continue

    print(f"[>] Analyzing {filename} in chunks...")

    # === Step 1: Prepare chunks and prompts
    chunks = list(chunk_code(asm_code, CHUNK_SIZE))
    prompts = [
        "Analyze this assembly code and tell me if it is malicious or benign. Explain your reasoning.\n\n"
        f"```\n{chunk}\n```"
        for chunk in chunks
    ]

    try:
        outputs = pipe(prompts, max_new_tokens=256, do_sample=True, temperature=0.2, batch_size=8)
    except Exception as e:
        print(f"[!] Error during batched inference: {e}")
        continue

    total_chunks = len(chunks)
    malicious_chunks = 0
    summaries = []
    labels = []

    for i, output in enumerate(outputs):
        result_text = output[0]["generated_text"]

        # Keep only the first '### Label:'
        if "### Label:" in result_text:
            result_text = result_text.split("### Label:")[0].strip() + "\n### Label: " + result_text.split("### Label:")[1].strip().split('\n')[0]

        label = extract_first_label(result_text)
        labels.append(label)

        if label == "malicious":
            malicious_chunks += 1

        # Extract the summary portion (everything before the label)
        summary_part = result_text.split("### Label:")[0].strip() if "### Label:" in result_text else result_text.strip()
        summaries.append(summary_part)

        print(f"  → Chunk {i+1}: {label}")

    # === Step 2: Voting threshold
    if total_chunks == 0:
        final_label = "undetermined"
    else:
        malicious_ratio = malicious_chunks / total_chunks
        final_label = "malicious" if malicious_ratio >= MALICIOUS_THRESHOLD else "benign"

    # === Step 3: Logging and Excel write
    short_asm = sanitize(asm_code[:300] + "..." if len(asm_code) > 300 else asm_code)
    result_summary = (
        f"Chunks: {total_chunks}, Malicious: {malicious_chunks} "
        f"→ Final Label: {final_label} ({round(malicious_ratio*100, 1)}%)"
    )

    combined_summaries = sanitize("\n---\n".join(summaries))
    ws.append([filename, short_asm, result_summary, combined_summaries])
    wb.save(excel_file)
    print(f"[✓] Saved result for {filename}: {result_summary}")


print(f"\n✅ All results written to: {excel_file}")
