from peft import PeftModel
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig, pipeline
import torch
import os
from glob import glob
from transformers import pipeline
from openpyxl import Workbook, load_workbook
import re
from langchain_community.vectorstores import FAISS
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain.embeddings.base import Embeddings
from transformers import AutoTokenizer, AutoModel

if torch.cuda.is_available():
    torch.cuda.empty_cache()

class CodeT5pEmbedding(Embeddings):
    def __init__(self, model_name="Salesforce/codet5p-220m"):
        self.tokenizer = AutoTokenizer.from_pretrained(model_name)
        self.model = AutoModel.from_pretrained(model_name)
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.model.to(self.device)
        self.model.eval()

    def embed_documents(self, texts):
        return [self._embed(text) for text in texts]

    def embed_query(self, text):
        return self._embed(text)

    def _embed(self, text):
        with torch.no_grad():
            tokens = self.tokenizer(text, return_tensors="pt", truncation=True, max_length=512, padding="max_length")
            tokens = {k: v.to(self.device) for k, v in tokens.items()}
            output = self.model.encoder(**tokens).last_hidden_state
            return output.mean(dim=1).squeeze().cpu().tolist()

# === Load vectorstore ===
embedding_model = CodeT5pEmbedding(model_name="Salesforce/codet5p-220m")
vectorstore = FAISS.load_local("malware_index_code5tE", embedding_model, allow_dangerous_deserialization=True)

# === System info ===
print(torch.__version__)
print(torch.cuda.is_available())
print("CUDA version (build):", torch.version.cuda)
print(torch.cuda.get_device_name(0))

# === Helpers ===
def sanitize(text):
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

def extract_first_label(generated_text):
    match = re.search(r"(?i)label\s*:\s*(malicious|benign)", generated_text)
    return match.group(1).lower() if match else "unknown"

# === Config ===
input_dir = "./input_mixed"
excel_file = "analysis_resultsCleanRAG_fullfileCode5t.xlsx"

# === BitsAndBytes config ===
bnb = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_use_double_quant=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_compute_dtype=torch.float16
)

# === Load model and tokenizer ===
base_model = AutoModelForCausalLM.from_pretrained(
    r"D:\MistraGood\Mistra",
    device_map="auto",
    quantization_config=bnb,
    trust_remote_code=True
)

model = base_model  # or load LoRA: PeftModel.from_pretrained(...)
tokenizer = AutoTokenizer.from_pretrained(r"D:\MistraGood\Mistra", use_fast=False)
tokenizer.pad_token = tokenizer.pad_token or tokenizer.eos_token
tokenizer.padding_side = "left"
pipe = pipeline("text-generation", model=model, tokenizer=tokenizer)

# === Excel setup ===
if os.path.exists(excel_file):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        print(f"[!] Corrupted Excel file. Re-creating it. Reason: {e}")
        os.remove(excel_file)
        wb = Workbook()
        ws = wb.active
        ws.append(["Filename", "Assembly Code (truncated)", "Analysis Result", "Model Summary"])
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Filename", "Assembly Code (truncated)", "Analysis Result", "Model Summary"])

# === Process files ===
for filepath in sorted(glob(os.path.join(input_dir, "*.txt"))):
    filename = os.path.basename(filepath)
    with open(filepath, "r", encoding="utf-8") as f:
        asm_code = f.read().strip()

    if not asm_code:
        print(f"[!] Skipping empty file: {filename}")
        continue

    print(f"[>] Analyzing {filename} as a full input...")

    # === Retrieve context via RAG
    query_vector = asm_code[:256]
    results = vectorstore.similarity_search(query_vector, k=3)
    rag_context = "\n\n".join([doc.page_content for doc in results])

    # === Build prompt
    prompt = f"""You are a malware analyst.
Here is assembly code data explained.
{rag_context}
Analyze the following full assembly code and tell me if it is malicious or benign. Justify your answer.

### Full Assembly Code:
{asm_code}

### Summary:"""

    try:
        output = pipe(prompt, max_new_tokens=512, do_sample=True, temperature=0.3)[0]["generated_text"]
    except Exception as e:
        print(f"[!] Inference error: {e}")
        continue

    if "### Label:" in output:
        result_text = output.split("### Label:")[0].strip() + "\n### Label: " + output.split("### Label:")[1].strip().split('\n')[0]
    else:
        result_text = output

    label = extract_first_label(result_text)

    short_asm = sanitize(asm_code[:300] + "..." if len(asm_code) > 300 else asm_code)
    combined_summary = sanitize(result_text)

    result_summary = f"Final Label: {label}"

    ws.append([filename, short_asm, result_summary, combined_summary])
    wb.save(excel_file)
    print(f"[✓] Saved result for {filename}: {label}")

print(f"\n✅ All results written to: {excel_file}")
